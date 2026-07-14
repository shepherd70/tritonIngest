"""Create privacy-safe cell digests with an independent XLSX reader."""

import argparse
import datetime as dt
import hashlib
import json
import math
import re
import struct

import openpyxl
from openpyxl.utils.datetime import to_excel


_NUMBER = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][+-]?\d+)?$")


def _double_token(value):
    value = float(value)
    if not math.isfinite(value):
        raise ValueError("Non-finite numeric workbook value")
    return "D" + struct.pack(">d", value).hex()


def _semantic_token(value, epoch):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return "B1" if value else "B0"
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return _double_token(to_excel(value, epoch))
    if isinstance(value, (int, float)):
        return _double_token(value)
    value = str(value).replace("\r\n", "\n").replace("\r", "\n")
    if value == "TRUE":
        return "B1"
    if value == "FALSE":
        return "B0"
    if _NUMBER.fullmatch(value):
        return _double_token(value)
    return "S" + value


def _sheet_digest(sheet, epoch):
    cells = []
    max_row = 0
    max_column = 0
    nonblank = 0
    for row in sheet.iter_rows():
        converted = []
        for cell in row:
            value = _semantic_token(cell.value, epoch)
            converted.append(value)
            if value is not None:
                nonblank += 1
                max_row = max(max_row, cell.row)
                max_column = max(max_column, cell.column)
        cells.append(converted)

    digest = hashlib.sha256()
    digest.update(f"{max_row}:{max_column}\n".encode())
    for row in range(max_row):
        for column in range(max_column):
            value = cells[row][column] if column < len(cells[row]) else None
            if value is None:
                digest.update(b"N\n")
            else:
                encoded = value.encode("utf-8")
                digest.update(f"V{len(encoded)}:".encode())
                digest.update(encoded)
                digest.update(b"\n")
    return {
        "rows": max_row,
        "columns": max_column,
        "nonblank": nonblank,
        "semantic_sha256": digest.hexdigest(),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("--sheet", action="append", required=True)
    args = parser.parse_args()

    source = hashlib.sha256()
    with open(args.workbook, "rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            source.update(block)

    workbook = openpyxl.load_workbook(
        args.workbook, read_only=True, data_only=True
    )
    sheets = []
    for name in args.sheet:
        if name not in workbook.sheetnames:
            raise SystemExit(f"Worksheet not found: {name}")
        result = _sheet_digest(workbook[name], workbook.epoch)
        result["name"] = name
        sheets.append(result)

    print(json.dumps({
        "schema_version": "triton-real-workbook-digest/v1",
        "engine": {"name": "openpyxl", "version": openpyxl.__version__},
        "source_sha256": source.hexdigest(),
        "sheets": sheets,
    }, ensure_ascii=False, separators=(",", ":")))


if __name__ == "__main__":
    main()
